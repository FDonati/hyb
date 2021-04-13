#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Fri Feb  5 14:17:17 2021

@author: franco
"""
from openpyxl import load_workbook
import pandas as pd
import numpy as np
import sqlite3

conn = sqlite3.connect("source_data/db.sqlite3")
cursor = conn.cursor()

E_path = "source_data/MR_HIOT_2011_v3_3_18_11_extensions.xlsx"
Y_path = "source_data/MR_HIOT_2011_v3_3_18_11_FD.csv"
Z_path = "source_data/MR_HIOT_2011_v3_3_18_11_by_product_technology.csv"
W_path = "source_data/MR_HIOT_2011_v3_3_18_11_stock_to_waste.csv"

inputToEconExt = ["va_act", "resource_", "use_", "waste_from"]

outputFromEconExt = ["emis", "sup_", "addition", "res"]

dataScheme = dict.fromkeys(["id", "value", "source", "target", "unit"])

def collectAllExt(path_data):

    workbook = load_workbook(path_data)

    sheet_names = workbook.sheetnames
    print(sheet_names)

    collection = {}
    for name in sheet_names:
        if name != "Sheet1":
            collection[name] = pd.read_excel(path_data, sheet_name=name, engine='openpyxl')

    return collection


def assembleExt(dictionary):

    Z_ext_labels = pd.DataFrame(columns=["name", "unit"])

    FD_ext_labels = Z_ext_labels.copy()

    Z_ext_array = np.empty([0, 7872], dtype=float)

    FD_ext_array = np.empty([0, 288], dtype=float)

    for key in dictionary:
        print(key.lower())

        if key != "VA_act":
            label = dictionary[key].iloc[3:, :2].reset_index(drop=True)
            label.columns = ["name", "unit"]
            ext_type = pd.DataFrame([key]*label.shape[0])
            ext_type.columns = ["ext_type"]
            label = pd.concat([ext_type, label], axis=1)
            array = dictionary[key].iloc[3:, 2:].to_numpy(dtype=float)

        else:
            label = dictionary[key].iloc[3:, :4].reset_index(drop=True)
            label.columns = ["code", "synonym", "name", "unit"]

            ext_type = pd.DataFrame([key]*label.shape[0])
            ext_type.columns = ["ext_type"]
            label = pd.concat(
                [ext_type, label.loc[:, ["name", "unit"]]], axis=1)
            array = dictionary[key].iloc[3:, 4:].to_numpy(dtype=float)

        if "fd" in key.lower() and key != "Avoided_emiss" or key == "waste_from_stocks":
            FD_ext_labels = FD_ext_labels.append(label, ignore_index=True)
            FD_ext_array = np.concatenate([FD_ext_array, array])
        else:
            Z_ext_labels = Z_ext_labels.append(label, ignore_index=True)
            Z_ext_array = np.concatenate([Z_ext_array, array])

    return {"Z_ext": {"Z_ext_labels": Z_ext_labels, "Z_ext_array": Z_ext_array},
            "FD_ext": {"FD_ext_labels": FD_ext_labels, "FD_ext_array": FD_ext_array}}


E_collect = collectAllExt(E_path)

E = assembleExt(E_collect)


Y = pd.read_csv(Y_path)
prod_labels = Y.iloc[3:, :5]
FD_labels = Y.iloc[:3, 5:].T.reset_index()
FD_array = np.array(Y.iloc[3:, 5:], dtype=float)
del Y

Z = pd.read_csv(Z_path)
act_labels = Z.iloc[:3, 5:].T.reset_index()
Z_array = np.array(Z.iloc[3:, 5:], dtype=float)
del Z


W = pd.read_csv(W_path)
W_array = np.array(W.iloc[:, 5:], dtype=float)


# Calculate Leontief inverse


def divideByOne(vector):
    mask = (vector == 0)
    vector[~mask] = 1/vector[~mask]

    return vector


x = FD_array.sum(axis=1) + Z_array.sum(axis=1)
x_diag_inv = divideByOne(np.diag(x))

A = Z_array @ x_diag_inv
I = np.identity(A.shape[0])

L = np.linalg.inv(I-A)

e = E["Z_ext"]["Z_ext_array"] @ x_diag_inv
w = W_array.T @ x_diag_inv

def save_arrays():
    np.save("data/L.npy", L)
    np.save("clean_data/hybrid/EY.npy", E["FD_ext"]["FD_ext_array"])
    np.save("clean_data/hybrid/e.npy", e)
    np.save("clean_data/hybrid/w.npy", w)

def save_labels():
    act_labels.to_csv("activities.csv")

# keywords contained in the labels of extension input


# For input values

# def forSankeyTest(prod_index, Y, Z):

#     selectedIndex = [num*163 + prod_index for num in range(48)]

#     i = np.zeros((7872), dtype=int)

#     collection = dict.fromkeys(range(48))

#     h = 0

#     for item in selectedIndex:

#         i[item] = 1

#         y = Y.sum(1) @ np.diag(i)

#         x = L @ y

#         E = e @ np.diag(x)

#         inputExt = dict.fromkeys(range(len(inputToEconExt)))
#         outputExt = dict.fromkeys(range(len(outputFromEconExt)))

#         data = {"x":x, "Y":Y[item],
#                 "Z_sup": Z[item], "Z_use": Z[:,item],
#                 "inputExt":E}

#         eIn = np.array_split(Ein, len(Ein))

#         EOut = eIn = np.array_split(Ein, len(Ein))

#         yOut = Y[item]

#         zOut = Z[:,item]


#         collection[h] = data

#         h += 1
#         if h == 0:
#             pass

#     return collection

# sankeyDataTest  = forSankeyTest(20, FD_array, Z_array)

# serializedSDT = json.JSONEncoder().iterencode(sankeyDataTest)

# with open("sankeyDataTest.json", "w") as outfile:
#     json.dump(sankeyDataTest, outfile)


# INPUTS
# Other activities => the economic input to an activity

# biosphere / resources
#   => to Output
#   => to Final Consumption

# Storage (reduction of inventory)

# OUTPUT

# Activities
#   => Stock addition
#   => Waste
#   => Unregistered waste
#   => Emissions
#       =>
#       =>
#       =>
#       =>
#       =>
#       =>
# Final Demand
#   => FD categories
#   => Stock addition
#   => Waste
#   => Emissions

# Previous accumulation of materials

# in
